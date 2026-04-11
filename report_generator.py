"""
report_generator.py — ThreatMap Infra Professional VAPT Report
Three sheets: Cover Page → Observations → Annexure

Root cause fix: sqlite3.Row objects converted to plain dicts at entry.
All subsequent access uses dict[] — no .get() calls on Row objects.
"""

import os
import glob
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ── Colour palette ───────────────────────────────────────────────────────────
BG_DARKEST  = "080C10"
BG_DARK     = "0D1117"
BG_MID      = "161B22"
BG_PANEL    = "1C2128"
BORDER_COL  = "30363D"
ACCENT_RED  = "E03B3B"
ACCENT_BLUE = "58A6FF"
WHITE       = "F0F6FC"
MUTED       = "8B949E"
FAINT       = "3D4450"

SEV_BG = {
    "Critical": "2D0000",
    "High":     "2D1200",
    "Medium":   "2D2200",
    "Low":      "0D2200",
    "Info":     "001830",
}
SEV_FG = {
    "Critical": "FF4C4C",
    "High":     "FF8C00",
    "Medium":   "E3B341",
    "Low":      "3FB950",
    "Info":     "58A6FF",
}
SEV_ORDER = ["Critical", "High", "Medium", "Low", "Info"]

# Every key that may appear in a triage record — with safe defaults
RECORD_DEFAULTS = {
    "host": "", "port": "", "service": "", "severity": "Info",
    "priority_rank": 5, "cvss_score": 0.0, "actively_exploited": 0,
    "observation_name": "", "detailed_observation": "",
    "impacted_module": "Network Service", "risk_impact": "",
    "risk_summary": "", "business_impact": "", "remediation": "",
    "false_positive_likelihood": "Low", "attack_scenario": "",
    "triage_method": "rule_based", "ai_enhanced": 0,
    "screenshot_path": None,
}


# ── Style helpers ────────────────────────────────────────────────────────────

def _fill(hex_col):
    return PatternFill("solid", fgColor=hex_col)

def _font(color=WHITE, bold=False, size=11, name="Arial"):
    return Font(color=color, bold=bold, size=size, name=name)

def _align(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def _border(color=BORDER_COL):
    s = Side(style="thin", color=color)
    return Border(left=s, right=s, top=s, bottom=s)

def _bg(ws, rows, cols, color):
    for r in range(1, rows + 1):
        for c in range(1, cols + 1):
            ws.cell(row=r, column=c).fill = _fill(color)


# ── Converter: sqlite3.Row (or any mapping) → clean dict ────────────────────

def _to_dict(row):
    """
    Convert sqlite3.Row to plain dict and fill missing keys with defaults.
    This is the ONLY place we touch raw DB rows — everything else uses dicts.
    """
    d = dict(RECORD_DEFAULTS)          # start with all defaults
    raw = dict(row)                     # sqlite3.Row supports dict()
    d.update({k: v for k, v in raw.items() if v is not None})
    # Normalise severity
    if d["severity"] not in SEV_ORDER:
        d["severity"] = "Info"
    return d


# ── Sheet 1: Cover Page ──────────────────────────────────────────────────────

def _build_cover(wb, records, scan_meta):
    ws = wb.active
    ws.title = "Cover Page"
    ws.sheet_view.showGridLines = False

    _bg(ws, 60, 14, BG_DARKEST)

    widths = [3, 20, 4, 30, 4, 14, 4, 14, 4, 3, 3, 3, 3, 3]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ── Red accent bars ──────────────────────────────────────────────
    for row in [1, 2]:
        for c in range(1, 15):
            ws.cell(row=row, column=c).fill = _fill(ACCENT_RED)
        ws.row_dimensions[row].height = 5

    # ── THREATMAP heading ────────────────────────────────────────────
    ws.row_dimensions[4].height = 52
    c = ws.cell(row=4, column=2, value="THREATMAP")
    c.font      = Font(name="Arial", bold=True, size=40, color=ACCENT_RED)
    c.alignment = _align("left", "center")
    ws.merge_cells("B4:I4")

    ws.row_dimensions[5].height = 22
    c = ws.cell(row=5, column=2,
                value="INFRA  ─  Vulnerability Assessment & Penetration Testing Report")
    c.font      = Font(name="Arial", size=13, color=MUTED)
    c.alignment = _align("left", "center")
    ws.merge_cells("B5:I5")

    # Thin red divider
    ws.row_dimensions[6].height = 3
    for col in range(2, 10):
        ws.cell(row=6, column=col).fill = _fill(ACCENT_RED)

    # ── Metadata block ───────────────────────────────────────────────
    target = scan_meta.get("target") or (records[0]["host"] if records else "—")
    date   = scan_meta.get("date") or datetime.now().strftime("%d %B %Y")
    mode   = (scan_meta.get("mode") or "balanced").title()

    meta_rows = [
        (8,  "TARGET",         target),
        (9,  "REPORT DATE",    date),
        (10, "SCAN MODE",      mode),
        (11, "CLASSIFICATION", "CONFIDENTIAL — For Authorized Recipients Only"),
        (12, "PREPARED BY",    "ThreatMap Infra v1.0 — Automated VAPT Engine"),
    ]
    for row, label, value in meta_rows:
        ws.row_dimensions[row].height = 22
        lc = ws.cell(row=row, column=2, value=label)
        lc.font      = Font(name="Arial", bold=True, size=9, color=MUTED)
        lc.alignment = _align("left", "center")
        vc = ws.cell(row=row, column=4, value=value)
        vc.font      = Font(name="Arial", size=10, color=WHITE)
        vc.alignment = _align("left", "center")
        ws.merge_cells(f"D{row}:I{row}")

    # ── Severity count boxes ─────────────────────────────────────────
    counts = {s: 0 for s in SEV_ORDER}
    for r in records:
        counts[r["severity"]] = counts.get(r["severity"], 0) + 1

    ws.row_dimensions[15].height = 10
    ws.row_dimensions[16].height = 44
    ws.row_dimensions[17].height = 20

    col = 2
    for sev in SEV_ORDER:
        sbg, sfg = SEV_BG[sev], SEV_FG[sev]
        cnt = counts.get(sev, 0)

        c = ws.cell(row=16, column=col, value=str(cnt))
        c.font      = Font(name="Arial", bold=True, size=24, color=sfg)
        c.fill      = _fill(sbg)
        c.alignment = _align("center", "center")
        ws.merge_cells(
            f"{get_column_letter(col)}16:{get_column_letter(col+1)}16"
        )
        c = ws.cell(row=17, column=col, value=sev.upper())
        c.font      = Font(name="Arial", bold=True, size=8, color=sfg)
        c.fill      = _fill(sbg)
        c.alignment = _align("center", "center")
        ws.merge_cells(
            f"{get_column_letter(col)}17:{get_column_letter(col+1)}17"
        )
        col += 2

    # ── Scope note ───────────────────────────────────────────────────
    ws.row_dimensions[20].height = 16
    c = ws.cell(row=20, column=2, value="SCOPE & METHODOLOGY")
    c.font = Font(name="Arial", bold=True, size=9, color=MUTED)

    scope = (
        f"This assessment was conducted against {target} using automated "
        "reconnaissance and vulnerability scanning tools including Nmap, Nikto, "
        "Gobuster, SSLScan, WhatWeb, and Nuclei. Findings are classified according "
        "to CVSS v3.1 and prioritised by exploitability and business impact."
    )
    ws.row_dimensions[21].height = 52
    c = ws.cell(row=21, column=2, value=scope)
    c.font      = Font(name="Arial", size=9, color=MUTED)
    c.alignment = _align("left", "top", wrap=True)
    ws.merge_cells("B21:I21")

    # ── Footer ───────────────────────────────────────────────────────
    for col in range(1, 15):
        ws.cell(row=56, column=col).fill = _fill(ACCENT_RED)
    ws.row_dimensions[56].height = 5
    ws.row_dimensions[57].height = 20
    c = ws.cell(
        row=57, column=2,
        value="CONFIDENTIAL — This report contains sensitive security information. "
              "Not for distribution beyond authorised recipients.",
    )
    c.font      = Font(name="Arial", size=8, color=WHITE)
    c.alignment = _align("center")
    ws.merge_cells("B57:I57")


# ── Sheet 2: Observations ────────────────────────────────────────────────────

def _build_observations(wb, records):
    ws = wb.create_sheet("Observations")
    ws.sheet_view.showGridLines = False

    _bg(ws, 500, 12, BG_DARK)

    col_widths = {
        "A": 3, "B": 7,  "C": 32, "D": 42,
        "E": 22, "F": 14, "G": 40, "H": 44, "I": 3,
    }
    for col, w in col_widths.items():
        ws.column_dimensions[col].width = w

    # Title bar
    ws.row_dimensions[1].height = 5
    for c in range(1, 10):
        ws.cell(row=1, column=c).fill = _fill(ACCENT_RED)

    ws.row_dimensions[2].height = 34
    c = ws.cell(row=2, column=2, value="OBSERVATIONS")
    c.font      = Font(name="Arial", bold=True, size=20, color=ACCENT_RED)
    c.alignment = _align("left", "center")

    # Summary line
    counts = {s: 0 for s in SEV_ORDER}
    for r in records:
        counts[r["severity"]] = counts.get(r["severity"], 0) + 1
    parts = [f"{counts[s]} {s}" for s in SEV_ORDER if counts.get(s, 0) > 0]

    ws.row_dimensions[3].height = 5
    ws.row_dimensions[4].height = 18
    c = ws.cell(
        row=4, column=2,
        value=f"Total: {len(records)}   ·   " + "   ·   ".join(parts)
    )
    c.font      = Font(name="Arial", size=9, color=MUTED)
    c.alignment = _align("left", "center")
    ws.merge_cells("B4:H4")

    ws.row_dimensions[5].height = 5

    # Column headers
    HEADERS = [
        (2, "S. No",                "center"),
        (3, "Observation Name",     "left"),
        (4, "Detailed Observation", "left"),
        (5, "Impacted Module",      "center"),
        (6, "Severity",             "center"),
        (7, "Risk / Impact",        "left"),
        (8, "Recommendation",       "left"),
    ]
    ws.row_dimensions[6].height = 26
    for col, header, align in HEADERS:
        c = ws.cell(row=6, column=col, value=header)
        c.fill      = _fill(BG_DARKEST)
        c.font      = Font(name="Arial", bold=True, size=10, color=ACCENT_BLUE)
        c.alignment = _align(align, "center")
        c.border    = _border()

    ws.freeze_panes = "B7"

    # Sort: priority ASC, cvss DESC
    sorted_records = sorted(
        records,
        key=lambda r: (int(r["priority_rank"]), -(float(r["cvss_score"])))
    )

    for i, rec in enumerate(sorted_records, 1):
        row    = 6 + i
        ws.row_dimensions[row].height = 72

        sev    = rec["severity"]
        sbg    = SEV_BG.get(sev, BG_MID)
        sfg    = SEV_FG.get(sev, MUTED)
        is_hot = sev in ("Critical", "High")
        row_bg = BG_PANEL if i % 2 == 0 else BG_MID

        # Derived fields with safe fallbacks
        obs = (rec["observation_name"]
               or f"Exposed {(rec['service'] or 'Unknown').upper()} Service")
        detail = (rec["detailed_observation"]
                  or rec["risk_summary"]
                  or f"Port {rec['port']}/TCP ({(rec['service'] or '').upper()}) is publicly accessible.")
        mod    = rec["impacted_module"] or "Network Service"
        risk   = rec["risk_impact"] or rec["business_impact"] or "Impact not assessed."
        fix    = rec["remediation"] or "Refer to vendor security guidance."

        # ── S. No ──
        c = ws.cell(row=row, column=2, value=i)
        c.fill      = _fill(sbg if is_hot else row_bg)
        c.font      = Font(name="Arial", bold=True, size=10,
                           color=sfg if is_hot else MUTED)
        c.alignment = _align("center", "top")
        c.border    = _border()

        # ── Observation Name ──
        c = ws.cell(row=row, column=3, value=obs)
        c.fill      = _fill(sbg if is_hot else row_bg)
        c.font      = Font(name="Arial", bold=True, size=10,
                           color=sfg if is_hot else WHITE)
        c.alignment = _align("left", "top", wrap=True)
        c.border    = _border()

        # ── Detailed Observation ──
        c = ws.cell(row=row, column=4, value=detail)
        c.fill      = _fill(row_bg)
        c.font      = Font(name="Arial", size=9, color=MUTED)
        c.alignment = _align("left", "top", wrap=True)
        c.border    = _border()

        # ── Impacted Module ──
        c = ws.cell(row=row, column=5, value=mod)
        c.fill      = _fill(row_bg)
        c.font      = Font(name="Arial", size=9, color=MUTED)
        c.alignment = _align("center", "top", wrap=True)
        c.border    = _border()

        # ── Severity ──
        c = ws.cell(row=row, column=6, value=sev)
        c.fill      = _fill(sbg)
        c.font      = Font(name="Arial", bold=True, size=10, color=sfg)
        c.alignment = _align("center", "center")
        c.border    = _border()

        # ── Risk / Impact ──
        c = ws.cell(row=row, column=7, value=risk)
        c.fill      = _fill(row_bg)
        c.font      = Font(name="Arial", size=9, color=MUTED)
        c.alignment = _align("left", "top", wrap=True)
        c.border    = _border()

        # ── Recommendation ──
        c = ws.cell(row=row, column=8, value=fix)
        c.fill      = _fill(row_bg)
        c.font      = Font(name="Arial", size=9, color=MUTED)
        c.alignment = _align("left", "top", wrap=True)
        c.border    = _border()


# ── Sheet 3: Annexure ────────────────────────────────────────────────────────

def _build_annexure(wb, scan_meta):
    ws = wb.create_sheet("Annexure")
    ws.sheet_view.showGridLines = False

    _bg(ws, 5000, 12, BG_DARK)

    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 24
    ws.column_dimensions["C"].width = 115
    ws.column_dimensions["D"].width = 3

    # Title bar
    ws.row_dimensions[1].height = 5
    for c in range(1, 12):
        ws.cell(row=1, column=c).fill = _fill(ACCENT_RED)

    ws.row_dimensions[2].height = 34
    c = ws.cell(row=2, column=2, value="ANNEXURE — SCAN EVIDENCE")
    c.font      = Font(name="Arial", bold=True, size=20, color=ACCENT_RED)
    c.alignment = _align("left", "center")

    ws.row_dimensions[3].height = 5
    ws.row_dimensions[4].height = 18
    c = ws.cell(row=4, column=2,
                value="Raw scan tool outputs captured during this assessment.")
    c.font      = Font(name="Arial", size=9, color=MUTED)
    c.alignment = _align("left", "center")

    row     = 6
    reports = "reports"
    target  = scan_meta.get("target", "target")

    SECTIONS = [
        ("NMAP — PORT SCAN",     f"nmap_{target}.xml"),
        ("WHOIS — REGISTRATION", f"whois_{target}.txt"),
        ("DNS — ENUMERATION",    f"dig_{target}.txt"),
        ("NIKTO — WEB SCAN",     f"nikto_{target}.txt"),
        ("GOBUSTER — DIRS",      f"gobuster_{target}.txt"),
        ("SSLSCAN — TLS",        f"sslscan_{target}.txt"),
        ("NUCLEI — CVE SCAN",    f"nuclei_{target}.txt"),
        ("CURL — HTTP HEADERS",  f"curl_headers_{target}.txt"),
        ("WHATWEB — TECH STACK", f"whatweb_{target}.json"),
        ("AUTHORIZATION LOG",    "authorization_log.txt"),
        ("SCAN LOG",             "scan.log"),
    ]

    for ef in glob.glob(f"{reports}/evidence_*.json"):
        SECTIONS.append((f"HTTP EVIDENCE — {Path(ef).name}", Path(ef).name))

    written = 0

    for title, filename in SECTIONS:
        fpath = os.path.join(reports, filename)

        # Wildcard fallback
        if not os.path.isfile(fpath):
            matches = glob.glob(
                os.path.join(reports, filename.replace(target, "*"))
            )
            if matches:
                fpath = matches[0]

        if not os.path.isfile(fpath):
            continue

        try:
            content = Path(fpath).read_text(
                encoding="utf-8", errors="replace"
            ).strip()
        except Exception:
            continue

        if not content:
            continue

        written += 1

        # Section header
        ws.row_dimensions[row].height = 26
        c = ws.cell(row=row, column=2, value=title)
        c.fill      = _fill(BG_DARKEST)
        c.font      = Font(name="Arial", bold=True, size=11, color=ACCENT_BLUE)
        c.alignment = _align("left", "center")
        c.border    = _border()
        ws.merge_cells(f"B{row}:C{row}")
        row += 1

        # Source path
        ws.row_dimensions[row].height = 14
        c = ws.cell(row=row, column=2, value=f"Source: {fpath}")
        c.fill      = _fill(BG_PANEL)
        c.font      = Font(name="Arial", size=8, color=FAINT)
        c.alignment = _align("left", "center")
        ws.merge_cells(f"B{row}:C{row}")
        row += 1

        # Content (cap 300 lines per section)
        lines = content.splitlines()
        if len(lines) > 300:
            lines = lines[:300]
            lines.append(f"[truncated — full output at {fpath}]")

        for line in lines:
            ws.row_dimensions[row].height = 13
            c = ws.cell(row=row, column=2, value=line)
            c.fill      = _fill(BG_MID if row % 2 == 0 else BG_DARK)
            c.font      = Font(name="Courier New", size=8, color=MUTED)
            c.alignment = _align("left", "center")
            ws.merge_cells(f"B{row}:C{row}")
            row += 1

        # Spacer between sections
        ws.row_dimensions[row].height = 8
        row += 1

    if written == 0:
        ws.row_dimensions[row].height = 22
        c = ws.cell(
            row=row, column=2,
            value="No scan logs found in reports/. Run a scan first.",
        )
        c.font      = Font(name="Arial", size=10, color=MUTED)
        c.alignment = _align("left", "center")


# ── Public API ───────────────────────────────────────────────────────────────

def generate_excel(db, output_path="reports/threatmap_report.xlsx"):
    """
    Build and save the three-sheet VAPT report.
    Converts sqlite3.Row objects to dicts before any processing.
    """
    raw = db.get_all_triage()

    if not raw:
        print("[!] No triage findings to export.")
        return None

    # ── THE FIX: convert every Row to a plain dict right here ─────────
    records = [_to_dict(r) for r in raw]

    # Scan metadata from DB
    scan_meta = {}
    try:
        with db._conn() as conn:
            row = conn.execute(
                "SELECT target, scan_mode, started_at, completed_at "
                "FROM scans ORDER BY id DESC LIMIT 1"
            ).fetchone()
            if row:
                scan_meta = {
                    "target": row["target"],
                    "mode":   row["scan_mode"],
                    "date":   (row["completed_at"] or row["started_at"] or "")[:10],
                }
    except Exception:
        pass

    wb = openpyxl.Workbook()

    _build_cover(wb, records, scan_meta)
    _build_observations(wb, records)
    _build_annexure(wb, scan_meta)

    # Sheet tab colours
    wb["Cover Page"].sheet_properties.tabColor   = "E03B3B"
    wb["Observations"].sheet_properties.tabColor = "58A6FF"
    wb["Annexure"].sheet_properties.tabColor     = "3D4450"
    wb.active = wb["Cover Page"]

    wb.save(output_path)

    total  = len(records)
    hosts  = len({r["host"] for r in records})
    ai_pct = int(100 * sum(1 for r in records if r["ai_enhanced"]) / max(total, 1))
    counts = {s: 0 for s in SEV_ORDER}
    for r in records:
        counts[r["severity"]] = counts.get(r["severity"], 0) + 1

    print(f"[✔] Report → {output_path}")
    print(f"    {total} findings · {hosts} host(s) · {ai_pct}% AI-enhanced")
    sev_str = "  ".join(f"{s}:{counts[s]}" for s in SEV_ORDER if counts.get(s))
    if sev_str:
        print(f"    {sev_str}")

    return output_path
